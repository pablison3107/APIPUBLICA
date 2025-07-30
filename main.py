from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from typing import List
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # ajuste para seu domínio depois
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def ler_planilha_excel(caminho: str) -> List[dict]:
    wb = load_workbook(caminho)
    ws = wb.active
    cabecalho = [cell.value for cell in ws[1]]
    dados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {chave: valor for chave, valor in zip(cabecalho, row)}
        dados.append(item)

    return dados

@app.get("/")
def root():
    return {"mensagem": "API da planilha está rodando!"}

@app.get("/api/dados")
def get_dados():
    caminho = os.path.join(os.getcwd(), "planilha.xlsx")
    try:
        dados = ler_planilha_excel(caminho)
        return dados
    except Exception as e:
        return {"erro": str(e)}
